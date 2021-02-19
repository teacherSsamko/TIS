import os
import datetime
import csv
import re

from pymongo import MongoClient
from openpyxl import Workbook, load_workbook

BASE_DIR = os.path.dirname(os.path.realpath(__file__))

mongo = MongoClient("mongodb://localhost:27017")
db = mongo['aircode']
col_schedule = db['shopnt_schedule']

schedule_file = '/Users/ssamko/Documents/ssamko/aircode/에어코드 제공 데이터/쇼핑엔티_20191101_20201031_편성표.xlsx'

wb = load_workbook(schedule_file, read_only=True)
ws = wb['Sheet1']
"""
on_date: 방송일자
start_time: 방송시작
end_time: 방송종료
play_period: 방송시작~종료
main: 메인(Bool)
prod_id: 상품코드
prod_name: 상품명
prod_price: 판매가
prod_stock: 주문가능수량
prod_status: 판매구분
prod_ctg: 대분류
MD: MD명
supplier: 공급업체명
QA: QA결과
gift_stockout: 사은품품절여부(bool)

"""

i = 1
r = []
for row in ws.rows:
    if i < 5:
        i += 1
        continue
    for cell in row:
        if cell.value: r.append(cell.value)

    if i % 2 == 0:
        print(r)
        # insert data to DB here
        on_time = r[-4].split(" ~ ")
        print(on_time)
        data = {
            'on_date': r[1],
            'start_time': on_time[0],
            'end_time': on_time[1],
            'play_period': r[-4],
            'main': r[3] == 'O',
            'prod_id': r[4],
            'prod_name': r[5],
            'prod_price': int(re.sub(',', '', r[6])),
            'prod_stock': int(re.sub(',', '', r[-3])),
            'prod_status': r[7],
            'prod_ctg': r[-2],
            'MD': r[8],
            'supplier': r[-1],
            'QA': r[9],
            'gift_stockout': r[10] == 'Y',

        }
        print(data)
        col_schedule.insert_one(data)
        r = []
    # if i == 10:
    #     break
    i += 1



# for row in ws.iter_rows():
#     for cell in row:
#         print(cell.value)
#         break
#     break

purchase_dir = '/Users/ssamko/Documents/ssamko/aircode/에어코드 제공 데이터/쇼핑엔티_20191101_20201031_주문데이터'

