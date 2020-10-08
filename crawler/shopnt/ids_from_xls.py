"""
결제내역
"""

import os
import datetime

from openpyxl import Workbook, load_workbook

today = datetime.date.today()

BASE_DIR = os.path.dirname(os.path.realpath(__file__))

wb = load_workbook(os.path.join(BASE_DIR,'shopnt_aug_receipt.xlsx'), read_only=True)

print(wb.sheetnames[0])
ws = wb['Select torderpromo']
print(ws)
print(ws.cell(row=2, column=4).value)
# max_row = sheet.max_row - 1
# max_row = ws.max_row
# print(max_row)
# max_col = ws.max_column - 1

with open(os.path.join(BASE_DIR,'prod_id_list08.txt'), 'w') as f:
#     for row_id in range(2, max_row):
#         prod_id = ws.cell(row=row_id, column=2).value
#         print(prod_id)
#         f.write(f'{prod_id}\n')
    prod_ids = set()
    for row in ws.iter_rows():
        # 중복은 제거
        prod_id = row[3].value
        print(prod_id)
        prod_ids.add(prod_id)

    print(len(prod_ids))
    prod_ids = list(prod_ids)
    print(len(prod_ids))

    for prod_id in prod_ids:
        if prod_id == '상품코드':
            continue
        f.write(f'{prod_id}\n')